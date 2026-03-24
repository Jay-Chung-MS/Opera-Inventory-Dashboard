"""
메디필 Inventory Dashboard - 데이터 파이프라인
S&OP 파일 + 가용재고 + B2B Top SKU → inventory_data.json → Dashboard
"""

import pandas as pd
import numpy as np
import openpyxl
import json
import sys
import os
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')

# ── 파일 경로 설정 ──────────────────────────────────────────
SALES_JSON = r'C:\Users\user\AppData\Local\Temp\opsdb\data.json'
SOP_FILE = r'C:\Users\user\Downloads\260318_SI Value up P2E_발주 운영 파일 3월 3주차_r126(PE 공유용).xlsx'
INVENTORY_FILE = r'C:\Users\user\Downloads\가용재고조회_20260318.xlsx'
B2B_FILE = r'C:\Users\user\Downloads\B2B_SKU_분석_통합본_260312.xlsx'
OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_JSON = os.path.join(OUTPUT_DIR, 'inventory_data.json')

def safe_float(val):
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return 0.0 if np.isnan(val) else float(val)
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0

def safe_int(val):
    return int(safe_float(val))

print("=" * 50)
print("Inventory Dashboard 데이터 업데이트")
print("=" * 50)

# ════════════════════════════════════════════════════
# 1. SKU Master
# ════════════════════════════════════════════════════
print("\n[1/5] SKU Master 로딩...")
wb = openpyxl.load_workbook(SOP_FILE, data_only=True)
ws_sku = wb['SKU MASTER - v2']

sku_map = {}  # 신품번 → SKU info
old_to_new = {}  # 구품번 → 신품번

for row in ws_sku.iter_rows(min_row=3, max_row=ws_sku.max_row, values_only=True):
    if row[0] is None:
        continue
    new_code = str(row[2]).strip() if row[2] else ''
    old_code = str(row[3]).strip() if row[3] else ''
    if not new_code:
        continue

    sku_map[new_code] = {
        'code': new_code,
        'old_code': old_code,
        'brand': str(row[1]) if row[1] else '',
        'line': str(row[4]) if row[4] else '',
        'cat_l': str(row[5]) if row[5] else '',
        'cat_m': str(row[6]) if row[6] else '',
        'cat_s': str(row[7]) if row[7] else '',
        'mgmt': str(row[9]) if row[9] else '',
        'status': str(row[10]) if row[10] else '',
        'name': str(row[11]) if row[11] else '',
        'spec': str(row[13]) if row[13] else '',
        'lead_time': safe_int(row[16]) if row[16] else 45,
        'moq': safe_int(row[18]) if len(row) > 18 and row[18] else 0,
        'price': 0,
    }
    if old_code:
        old_to_new[old_code] = new_code

print(f"  {len(sku_map)}개 SKU, {len(old_to_new)}개 코드 매핑")

# ════════════════════════════════════════════════════
# 2. 가용재고
# ════════════════════════════════════════════════════
print("[2/5] 가용재고 로딩...")
df_inv = pd.read_excel(INVENTORY_FILE, header=1)
df_inv.columns = [str(c).strip() for c in df_inv.columns]
df_inv = df_inv[df_inv['품번'].notna() & df_inv['사업단위'].notna()]
# 원부자재 제거 → 제품만
if '품목자산분류' in df_inv.columns:
    before = len(df_inv)
    df_inv = df_inv[df_inv['품목자산분류'].astype(str).str.contains('제품', na=False)]
    print(f"  원부자재 제거: {before} → {len(df_inv)}행 (제품만)")
df_inv['품번'] = df_inv['품번'].astype(str).str.strip()
for col in ['현재고', '가용재고', '입고예정', '출고예정', '안전재고']:
    df_inv[col] = pd.to_numeric(df_inv[col], errors='coerce').fillna(0)

# 창고별 상세
warehouse_detail = {}
for _, r in df_inv.iterrows():
    code = r['품번']
    wh = str(r.get('창고', '기타'))
    if code not in warehouse_detail:
        warehouse_detail[code] = {}
    warehouse_detail[code][wh] = warehouse_detail[code].get(wh, 0) + int(r['현재고'])

# 품번별 합산
inv_agg = df_inv.groupby('품번').agg({
    '현재고': 'sum', '입고예정': 'sum', '출고예정': 'sum',
    '안전재고': 'sum', '가용재고': 'sum'
}).to_dict('index')

print(f"  {len(inv_agg)}개 품목")

# ════════════════════════════════════════════════════
# 3. 출고실적 + 판매계획
# ════════════════════════════════════════════════════
print("[3/5] 출고실적 + 판매계획...")

# 출고실적
ws_ship = wb['출고실적']
ship_labels = {6:'24-09',7:'24-10',8:'24-11',9:'24-12',
    10:'25-01',11:'25-02',12:'25-03',13:'25-04',14:'25-05',15:'25-06',
    16:'25-07',17:'25-08',18:'25-09',19:'25-10',20:'25-11',21:'25-12',
    22:'26-01',23:'26-02'}

ship_by_sku = {}  # {신품번: {월: qty}}
for row in ws_ship.iter_rows(min_row=7, max_row=ws_ship.max_row, values_only=True):
    new_code = str(row[1]).strip() if row[1] else '-'
    old_code = str(row[2]).strip() if row[2] else ''
    ship_type = str(row[4]) if row[4] else ''

    if new_code == '-' and old_code:
        new_code = old_to_new.get(old_code, '-')

    if new_code == '-' or '정상' not in ship_type:
        continue

    if new_code not in ship_by_sku:
        ship_by_sku[new_code] = {}

    for col_idx, label in ship_labels.items():
        val = safe_float(row[col_idx - 1])
        ship_by_sku[new_code][label] = ship_by_sku[new_code].get(label, 0) + val

# 판매계획
ws_plan = wb['판매계획']
plan_orig = {5:'25-04',6:'25-05',7:'25-06',8:'25-07',9:'25-08',10:'25-09',
    11:'25-10',12:'25-11',13:'25-12',14:'26-01',15:'26-02',16:'26-03',
    17:'26-04',18:'26-05',19:'26-06',20:'26-07'}
plan_changed = {22:'25-05',23:'25-06',24:'25-07',25:'25-08',26:'25-09',
    27:'25-10',28:'25-11',29:'25-12',30:'26-01',31:'26-02',32:'26-03',
    33:'26-04',34:'26-05',35:'26-06',36:'26-07'}

plan_by_sku = {}
for row in ws_plan.iter_rows(min_row=5, max_row=ws_plan.max_row, values_only=True):
    code = str(row[1]).strip() if row[1] else ''
    if not code:
        continue

    plan_by_sku[code] = {}
    # 원래 계획
    for col_idx, label in plan_orig.items():
        plan_by_sku[code][label] = safe_float(row[col_idx - 1])
    # 변경적용 (0이 아니면 덮어쓰기)
    for col_idx, label in plan_changed.items():
        val = safe_float(row[col_idx - 1])
        if val != 0:
            plan_by_sku[code][label] = val

wb.close()
print(f"  출고: {len(ship_by_sku)}개 SKU, 판매계획: {len(plan_by_sku)}개")

# ════════════════════════════════════════════════════
# 4. B2B Top SKU
# ════════════════════════════════════════════════════
print("[4/5] B2B Top SKU...")
wb_b2b = openpyxl.load_workbook(B2B_FILE, data_only=True)

b2b_channels = {}

overseas_tabs = {'해외_이에스_SKU':'이에스(CIS)','해외_동아_SKU':'동아(중국/동남아)',
    '해외_에이앤리_SKU':'에이앤리(유럽)','해외_힐락_SKU':'힐락(일본)'}
domestic_tabs = {'다이소 Top30':'다이소','올리브영 Top30':'올리브영','쿠팡 Top30':'쿠팡'}

for tab, ch in overseas_tabs.items():
    ws = wb_b2b[tab]
    items = []
    for row in ws.iter_rows(min_row=5, max_row=min(25, ws.max_row), values_only=True):
        if row[0] and row[1]:
            old = str(row[1]).strip()
            new = old_to_new.get(old, '')
            items.append({
                'rank': int(row[0]) if isinstance(row[0], (int, float)) else 0,
                'old_code': old, 'code': new,
                'name': str(row[2])[:55] if row[2] else '',
                'qty': safe_int(row[4]), 'amount': safe_int(row[5]),
            })
    b2b_channels[ch] = items

for tab, ch in domestic_tabs.items():
    ws = wb_b2b[tab]
    items = []
    for row in ws.iter_rows(min_row=4, max_row=min(34, ws.max_row), values_only=True):
        if row[0] and row[1]:
            code = str(row[1]).strip()
            items.append({
                'rank': int(row[0]) if isinstance(row[0], (int, float)) else 0,
                'old_code': '', 'code': code,
                'name': str(row[2])[:55] if row[2] else '',
                'qty': safe_int(row[6]), 'amount': safe_int(row[5]),
            })
    b2b_channels[ch] = items

wb_b2b.close()

# 전체 B2B Top SKU 신품번 집합
b2b_codes = set()
for ch, items in b2b_channels.items():
    for it in items:
        if it['code']:
            b2b_codes.add(it['code'])

print(f"  {len(b2b_channels)}개 채널, 고유 SKU: {len(b2b_codes)}개")

# ════════════════════════════════════════════════════
# 5. 분석 + JSON 생성
# ════════════════════════════════════════════════════
print("[5/5] 분석 & JSON 생성...")

# ── 수요예측 ──
use_months = sorted([m for m in ship_labels.values() if m >= '25-01'])

forecasts = {}
for code in sku_map:
    if code not in ship_by_sku:
        continue

    vals = [ship_by_sku[code].get(m, 0) for m in use_months]
    v3 = vals[-3:] if len(vals) >= 3 else vals
    v6 = vals[-6:] if len(vals) >= 6 else vals

    if sum(v3) == 0:
        continue

    sma3 = np.mean(v3)
    wma3 = (v3[0]*1 + v3[1]*2 + v3[2]*3) / 6 if len(v3) == 3 else sma3
    ets = v6[0]
    for v in v6[1:]:
        ets = 0.3 * v + 0.7 * ets

    if len(v6) >= 3:
        coeffs = np.polyfit(np.arange(len(v6)), v6, 1)
        slope = coeffs[0]
        linear = [max(0, coeffs[1] + slope * (len(v6) + i)) for i in range(3)]
    else:
        slope = 0
        linear = [sma3] * 3

    recommended = (wma3 + ets) / 2
    cv = np.std(v6) / np.mean(v6) * 100 if np.mean(v6) > 0 else 0

    forecasts[code] = {
        'sma3': round(sma3), 'wma3': round(wma3), 'ets': round(ets),
        'recommended': round(recommended),
        'linear': [round(x) for x in linear],
        'slope': round(slope, 1), 'cv': round(cv, 1),
        'monthly': {m: round(ship_by_sku[code].get(m, 0)) for m in use_months},
    }

# ── 재고 알람 ──
alert_list = []
alert_counts = {'critical': 0, 'warning': 0, 'caution': 0, 'normal': 0,
                'surplus': 0, 'excess': 0, 'slow': 0, 'none': 0}

for code, sku in sku_map.items():
    if '단종' in sku['status']:
        continue

    inv = inv_agg.get(code, {})
    available = int(inv.get('가용재고', 0))
    current = int(inv.get('현재고', 0))
    incoming = int(inv.get('입고예정', 0))
    outgoing = int(inv.get('출고예정', 0))

    fc = forecasts.get(code, {})
    rec = fc.get('recommended', 0)
    daily = rec / 30 if rec > 0 else 0
    days_left = round(available / daily) if daily > 0 else 9999
    lt = sku['lead_time'] if sku['lead_time'] > 0 else 45

    if rec == 0:
        if available > 0:
            level = 'slow'
        else:
            level = 'none'
    elif days_left <= lt * 0.5:
        level = 'critical'
    elif days_left <= lt:
        level = 'warning'
    elif days_left <= lt * 1.5:
        level = 'caution'
    elif days_left >= lt * 6:
        level = 'excess'
    elif days_left >= lt * 4:
        level = 'surplus'
    else:
        level = 'normal'

    alert_counts[level] = alert_counts.get(level, 0) + 1

    # 채널 태그
    channels = []
    for ch, items in b2b_channels.items():
        for it in items:
            if it['code'] == code:
                channels.append(ch)
                break

    # 판매계획
    plan = plan_by_sku.get(code, {})
    plan_26_03 = plan.get('26-03', 0)
    plan_26_04 = plan.get('26-04', 0)

    entry = {
        'code': code, 'name': sku['name'][:45],
        'line': sku['line'], 'mgmt': sku['mgmt'], 'status': sku['status'],
        'lead_time': lt,
        'current': current, 'available': available,
        'incoming': incoming, 'outgoing': outgoing,
        'forecast': rec, 'daily': round(daily, 1),
        'days_left': days_left if days_left < 9999 else None,
        'level': level,
        'channels': channels,
        'plan_03': round(plan_26_03), 'plan_04': round(plan_26_04),
        'sma3': fc.get('sma3', 0), 'wma3': fc.get('wma3', 0), 'ets': fc.get('ets', 0),
        'cv': fc.get('cv', 0), 'slope': fc.get('slope', 0),
        'is_b2b_top': code in b2b_codes,
        # SKU 360°: 월별 출고 트렌드 + 판매계획
        'ship_months': use_months,
        'ship_trend': [round(ship_by_sku.get(code, {}).get(m, 0)) for m in use_months],
        'plan_months': sorted(plan_orig.values()),
        'plan_trend': [round(plan_by_sku.get(code, {}).get(m, 0)) for m in sorted(plan_orig.values())],
        # 채널별 공급액
        'ch_amounts': {ch: next((it['amount'] for it in items if it['code'] == code), 0)
                       for ch, items in b2b_channels.items()},
        # 추가 SKU 정보
        'spec': sku['spec'], 'moq': sku['moq'], 'cat_m': sku['cat_m'], 'cat_s': sku['cat_s'],
    }
    alert_list.append(entry)

# 정렬: 알람 레벨 우선 → 재고소진일수 오름차순
level_order = {'critical': 0, 'warning': 1, 'caution': 2, 'normal': 3,
               'surplus': 4, 'excess': 5, 'slow': 6, 'none': 7}
alert_list.sort(key=lambda x: (level_order.get(x['level'], 9),
                                x['days_left'] if x['days_left'] else 9999))

# ── FCST 정확도 ──
common_months = sorted(set(ship_labels.values()) & set(plan_orig.values()))
fcst_monthly = []
for m in common_months:
    total_plan = sum(plan_by_sku.get(c, {}).get(m, 0) for c in plan_by_sku)
    total_actual = sum(ship_by_sku.get(c, {}).get(m, 0) for c in ship_by_sku)
    if total_plan > 0:
        dev = (total_actual - total_plan) / total_plan * 100
        acc = max(0, 100 - abs(dev))
    else:
        dev = 0
        acc = 0
    fcst_monthly.append({
        'month': m, 'plan': round(total_plan), 'actual': round(total_actual),
        'deviation': round(dev, 1), 'accuracy': round(acc, 1)
    })

# B2B Top SKU만 FCST
b2b_fcst_monthly = []
for m in common_months:
    total_plan = sum(plan_by_sku.get(c, {}).get(m, 0) for c in b2b_codes if c in plan_by_sku)
    total_actual = sum(ship_by_sku.get(c, {}).get(m, 0) for c in b2b_codes if c in ship_by_sku)
    if total_plan > 0:
        dev = (total_actual - total_plan) / total_plan * 100
        acc = max(0, 100 - abs(dev))
    else:
        dev = 0
        acc = 0
    b2b_fcst_monthly.append({
        'month': m, 'plan': round(total_plan), 'actual': round(total_actual),
        'deviation': round(dev, 1), 'accuracy': round(acc, 1)
    })

# ── 카테고리별 재고 분포 (알람별 stacked) ──
cat_inventory = {}
for entry in alert_list:
    cat = entry['line'] or '기타'
    if cat not in cat_inventory:
        cat_inventory[cat] = {'count': 0, 'available': 0, 'current': 0,
                              'alert_critical': 0, 'alert_risk': 0,
                              'alert_normal': 0, 'alert_excess': 0, 'alert_slow': 0}
    cat_inventory[cat]['count'] += 1
    cat_inventory[cat]['available'] += entry['available']
    cat_inventory[cat]['current'] += entry['current']
    lvl = entry['level']
    if lvl == 'critical':
        cat_inventory[cat]['alert_critical'] += 1
    elif lvl in ('warning', 'caution'):
        cat_inventory[cat]['alert_risk'] += 1
    elif lvl in ('normal', 'surplus'):
        cat_inventory[cat]['alert_normal'] += 1
    elif lvl == 'excess':
        cat_inventory[cat]['alert_excess'] += 1
    elif lvl == 'slow':
        cat_inventory[cat]['alert_slow'] += 1

# 상위 15개 라인만
cat_sorted = sorted(cat_inventory.items(), key=lambda x: x[1]['available'], reverse=True)[:15]

# ── 채널별 B2B 요약 ──
channel_summary = {}
for ch, items in b2b_channels.items():
    ch_data = {'name': ch, 'sku_count': len(items), 'total_amount': 0,
               'alerts': {'critical': 0, 'warning': 0, 'normal': 0, 'excess': 0},
               'skus': []}
    for it in items:
        ch_data['total_amount'] += it['amount']
        # 해당 SKU의 알람 정보 찾기
        code = it['code']
        entry = next((e for e in alert_list if e['code'] == code), None)
        if entry:
            lvl = entry['level']
            if lvl in ('critical', 'warning'):
                ch_data['alerts'][lvl] = ch_data['alerts'].get(lvl, 0) + 1
            elif lvl in ('excess', 'surplus', 'slow'):
                ch_data['alerts']['excess'] = ch_data['alerts'].get('excess', 0) + 1
            else:
                ch_data['alerts']['normal'] = ch_data['alerts'].get('normal', 0) + 1

            ch_data['skus'].append({
                'rank': it['rank'], 'code': code, 'old_code': it['old_code'],
                'name': it['name'][:45],
                'amount': it['amount'], 'qty': it['qty'],
                'available': entry['available'],
                'forecast': entry['forecast'],
                'days_left': entry['days_left'],
                'level': entry['level'],
            })
    channel_summary[ch] = ch_data

# ── 가용재고 파일 전체 총량 (SKU Master 무관) ──
inv_total_current = int(df_inv['현재고'].sum())
inv_total_available = int(df_inv['가용재고'].sum())
inv_total_incoming = int(df_inv['입고예정'].sum())
inv_total_outgoing = int(df_inv['출고예정'].sum())
inv_total_items = len(df_inv.groupby('품번'))

# ── SKU 평균 FCST 정확도 (개별 SKU 동일 가중치) ──
sku_acc_records = []
for code in set(ship_by_sku.keys()) & set(plan_by_sku.keys()):
    for m in common_months:
        p = plan_by_sku.get(code, {}).get(m, 0)
        a_val = ship_by_sku.get(code, {}).get(m, 0)
        if p > 0:
            sku_acc_records.append(max(0, 100 - abs((a_val - p) / p * 100)))

sku_avg_accuracy_all = round(np.mean(sku_acc_records), 1) if sku_acc_records else 0

sku_acc_b2b = []
for code in b2b_codes:
    if code not in ship_by_sku or code not in plan_by_sku:
        continue
    for m in common_months:
        p = plan_by_sku.get(code, {}).get(m, 0)
        a_val = ship_by_sku.get(code, {}).get(m, 0)
        if p > 0:
            sku_acc_b2b.append(max(0, 100 - abs((a_val - p) / p * 100)))

sku_avg_accuracy_b2b = round(np.mean(sku_acc_b2b), 1) if sku_acc_b2b else 0

# ── 채널별 SKU에 출고 트렌드 추가 ──
for ch, ch_data in channel_summary.items():
    for sku_entry in ch_data['skus']:
        code = sku_entry['code']
        fc = forecasts.get(code, {})
        monthly = fc.get('monthly', {})
        # 최근 6개월 출고 트렌드
        sku_entry['trend'] = [round(monthly.get(m, 0)) for m in use_months[-6:]]
        sku_entry['trend_months'] = use_months[-6:]

# ── S95/S90 과부족 + 입고계획 ──
print("[5-b] S95/S90 과부족 + 입고계획...")
wb_sop = openpyxl.load_workbook(SOP_FILE, data_only=True)

schedule_items = []

for tier_name, sheet_name in [('S95', 'MTS-s95'), ('S90', 'MTS-s90')]:
    try:
        ws_mts = wb_sop[sheet_name]
    except KeyError:
        print(f"  시트 '{sheet_name}' 없음 - 건너뜀")
        continue

    rows_data = list(ws_mts.iter_rows(min_row=14, max_row=ws_mts.max_row, values_only=True))
    if len(rows_data) < 3:
        continue

    # Row 14 (index 0): date headers from col 33 onwards (0-indexed col 32)
    date_row = rows_data[0]

    # Extract monthly history labels (cols 33-50, 0-indexed 32-49)
    history_months = []
    for ci in range(32, min(50, len(date_row))):
        val = date_row[ci]
        if val is None:
            break
        if hasattr(val, 'strftime'):
            history_months.append(val.strftime('%y-%m'))
        else:
            history_months.append(str(val))

    # Schedule period columns: 53,54=3월상/하, 55,56=4월상/하, ... 63,64=8월상/하
    schedule_periods = ['3월상', '3월하', '4월상', '4월하', '5월상', '5월하',
                        '6월상', '6월하', '7월상', '7월하', '8월상', '8월하']
    schedule_cols = [52, 53, 54, 55, 56, 57, 58, 59, 60, 61, 62, 63]  # 0-indexed

    # Data starts at row 17 (index 3 from row 14), each product uses 6 rows
    data_start_idx = 3  # row 17 = index 3 relative to row 14
    r_idx = data_start_idx
    while r_idx + 5 < len(rows_data):
        row_incoming = rows_data[r_idx]       # 입고
        row_outgoing = rows_data[r_idx + 1]   # 출고 Total
        row_balance = rows_data[r_idx + 5]    # 과부족

        # Validate this is a product block by checking type markers
        type_check_incoming = row_incoming[30] if len(row_incoming) > 30 else None
        type_check_balance = row_balance[30] if len(row_balance) > 30 else None

        # Extract product code from first row of block
        code = str(row_incoming[2]).strip() if row_incoming[2] else ''
        if not code or code == 'None':
            r_idx += 6
            continue

        name = str(row_incoming[7]).strip() if row_incoming[7] else ''
        line = str(row_incoming[6]).strip() if row_incoming[6] else ''
        spec = str(row_incoming[8]).strip() if row_incoming[8] else ''
        lead_time = safe_int(row_incoming[15]) if len(row_incoming) > 15 else 0
        avg_sales = safe_int(row_incoming[16]) if len(row_incoming) > 16 else 0
        current_stock = safe_int(row_incoming[17]) if len(row_incoming) > 17 else 0
        stock_months = safe_float(row_incoming[18]) if len(row_incoming) > 18 else 0
        safety_stock = safe_int(row_incoming[20]) if len(row_incoming) > 20 else 0
        days_after_lt = safe_float(row_incoming[22]) if len(row_incoming) > 22 else 0
        order_needed = str(row_incoming[23]).strip() if len(row_incoming) > 23 and row_incoming[23] else ''
        order_qty = safe_int(row_incoming[26]) if len(row_incoming) > 26 else 0
        issue_text = str(row_incoming[28]).strip() if len(row_incoming) > 28 and row_incoming[28] else ''

        # Monthly history values
        hist_incoming = []
        hist_outgoing = []
        hist_balance = []
        for ci in range(32, 32 + len(history_months)):
            if ci < len(row_incoming):
                hist_incoming.append(safe_int(row_incoming[ci]))
            else:
                hist_incoming.append(0)
            if ci < len(row_outgoing):
                hist_outgoing.append(safe_int(row_outgoing[ci]))
            else:
                hist_outgoing.append(0)
            if ci < len(row_balance):
                hist_balance.append(safe_int(row_balance[ci]))
            else:
                hist_balance.append(0)

        # Future schedule (상/하반월)
        schedule = []
        for pi, col_i in enumerate(schedule_cols):
            if pi < len(schedule_periods):
                inc_val = safe_int(row_incoming[col_i]) if col_i < len(row_incoming) else 0
                out_val = safe_int(row_outgoing[col_i]) if col_i < len(row_outgoing) else 0
                bal_val = safe_int(row_balance[col_i]) if col_i < len(row_balance) else 0
                schedule.append({
                    'period': schedule_periods[pi],
                    'incoming': inc_val,
                    'outgoing': out_val,
                    'balance': bal_val,
                })

        schedule_item = {
            'code': code,
            'name': name,
            'line': line,
            'tier': tier_name,
            'lt': lead_time,
            'avg_sales': avg_sales,
            'stock': current_stock,
            'stock_months': round(stock_months, 1),
            'safety_stock': safety_stock,
            'days_after_lt': round(days_after_lt, 1) if days_after_lt else 0,
            'order_needed': order_needed,
            'order_qty': order_qty,
            'issue': issue_text,
            'history_months': history_months,
            'history_incoming': hist_incoming,
            'history_outgoing': hist_outgoing,
            'history_balance': hist_balance,
            'schedule': schedule,
        }
        schedule_items.append(schedule_item)
        r_idx += 6

    print(f"  {tier_name}: {sum(1 for x in schedule_items if x['tier']==tier_name)}개 제품")

# Sort by days_after_lt ascending
schedule_items.sort(key=lambda x: x['days_after_lt'] if x['days_after_lt'] else 9999)

# 입고계획 (Pending Orders)
pending_orders = []
try:
    ws_po = wb_sop['입고계획']
    for row in ws_po.iter_rows(min_row=6, max_row=ws_po.max_row, values_only=True):
        if row[1] is None:
            continue
        status = str(row[9]).strip() if len(row) > 9 and row[9] else ''
        qty = safe_int(row[8]) if len(row) > 8 else 0
        if status == '입고완료' or qty <= 0:
            continue

        order_date_val = row[2] if len(row) > 2 else None
        if hasattr(order_date_val, 'strftime'):
            order_date_str = order_date_val.strftime('%Y-%m-%d')
        else:
            order_date_str = str(order_date_val) if order_date_val else ''

        period_val = str(row[7]).strip() if len(row) > 7 and row[7] else ''

        pending_orders.append({
            'order_no': str(row[1]).strip() if len(row) > 1 and row[1] else '',
            'order_date': order_date_str,
            'code': str(row[4]).strip() if len(row) > 4 and row[4] else '',
            'name': str(row[5]).strip() if len(row) > 5 and row[5] else '',
            'period': period_val,
            'qty': qty,
            'status': status,
        })
except KeyError:
    print("  시트 '입고계획' 없음")

# Sort pending orders by period
period_order = {'3상': 0, '3하': 1, '4상': 2, '4하': 3, '5상': 4, '5하': 5,
                '6상': 6, '6하': 7, '7상': 8, '7하': 9, '8상': 10, '8하': 11}
pending_orders.sort(key=lambda x: period_order.get(x['period'], 99))

s95_items = [x for x in schedule_items if x['tier'] == 'S95']
s90_items = [x for x in schedule_items if x['tier'] == 'S90']
order_needed_count = sum(1 for x in schedule_items if '발주필요' in x.get('order_needed', ''))

wb_sop.close()
print(f"  총 {len(schedule_items)}개 제품, 입고대기 {len(pending_orders)}건 ({sum(x['qty'] for x in pending_orders):,}개)")

# ── 이전 데이터 읽기 (주간 변화 추적) ──
prev_alerts = {}
prev_schedule = {}
try:
    with open(OUTPUT_JSON, encoding='utf-8') as f:
        prev_data = json.load(f)
    prev_alerts = prev_data.get('summary', {}).get('alerts', {})
    prev_schedule = prev_data.get('schedule', {}).get('summary', {})
    print(f"  이전 데이터 로드 완료: 긴급 {prev_alerts.get('critical', '?')}개")
except Exception as ex:
    print(f"  이전 데이터 없음 (첫 실행): {ex}")

# ── 입고 vs 판매계획 Gap (향후 3개월) ──
print("[5-c] 입고 vs 판매계획 Gap 분석...")
gap_items = []
for si in schedule_items:
    code = si['code']
    tier = si['tier']
    current_stock = si['stock']

    # fcst_3m: plan_by_sku에서 26-03, 26-04, 26-05 합산
    plan = plan_by_sku.get(code, {})
    fcst_3m = round(safe_float(plan.get('26-03', 0)) + safe_float(plan.get('26-04', 0)) + safe_float(plan.get('26-05', 0)))

    # fcst가 0이면 alerts에서 forecast 기반으로 대체
    if fcst_3m == 0:
        entry = next((e for e in alert_list if e['code'] == code), None)
        if entry and entry['forecast'] > 0:
            fcst_3m = round(entry['forecast'] * 3)

    if fcst_3m == 0:
        continue

    # incoming_3m: schedule에서 3월상~5월하 incoming 합산
    incoming_3m = 0
    target_periods = {'3월상', '3월하', '4월상', '4월하', '5월상', '5월하'}
    for s in si.get('schedule', []):
        if s['period'] in target_periods and s['incoming'] > 0:
            incoming_3m += s['incoming']

    gap = current_stock + incoming_3m - fcst_3m
    if gap < 0:
        gap_status = 'shortage'
    elif gap < fcst_3m * 0.3:
        gap_status = 'tight'
    else:
        gap_status = 'ok'

    name = si['name']
    gap_items.append({
        'code': code,
        'name': name,
        'tier': tier,
        'current_stock': current_stock,
        'fcst_3m': fcst_3m,
        'incoming_3m': incoming_3m,
        'gap': gap,
        'gap_status': gap_status,
    })

gap_items.sort(key=lambda x: x['gap'])
shortage_count = sum(1 for g in gap_items if g['gap_status'] == 'shortage')
tight_count = sum(1 for g in gap_items if g['gap_status'] == 'tight')
ok_count = sum(1 for g in gap_items if g['gap_status'] == 'ok')
total_gap = sum(g['gap'] for g in gap_items)
print(f"  Gap 분석: {len(gap_items)}개 제품 / 부족 {shortage_count} / 주의 {tight_count} / 정상 {ok_count}")

# ── 채널별 결품 위험 SKU 수 ──
for ch_name, ch_data in channel_summary.items():
    risk_skus = [s for s in ch_data['skus'] if s['level'] in ('critical', 'warning')]
    risk_skus.sort(key=lambda x: (x['days_left'] if x['days_left'] is not None else 9999))
    ch_data['risk_skus'] = len(risk_skus)
    ch_data['risk_sku_list'] = [{'code': s['code'], 'name': s['name'], 'level': s['level'], 'days_left': s['days_left']} for s in risk_skus[:5]]

# ── JSON 출력 ──
total_active = sum(1 for e in alert_list if e['level'] != 'none')

output = {
    'updated': datetime.now().strftime('%Y-%m-%d %H:%M'),
    'summary': {
        'total_sku': len(sku_map),
        'active_sku': total_active,
        # 가용재고 파일 전체 기준 (SKU Master 무관)
        'inv_total_current': inv_total_current,
        'inv_total_available': inv_total_available,
        'inv_total_incoming': inv_total_incoming,
        'inv_total_outgoing': inv_total_outgoing,
        'inv_total_items': inv_total_items,
        # 활성 SKU 기준
        'active_current': sum(e['current'] for e in alert_list),
        'active_available': sum(e['available'] for e in alert_list),
        'b2b_top_count': len(b2b_codes),
        'alerts': alert_counts,
    },
    'alerts': alert_list,
    'fcst': {
        'all': fcst_monthly,
        'b2b_top': b2b_fcst_monthly,
        # 총량 기준 정확도
        'overall_accuracy': round(np.mean([f['accuracy'] for f in fcst_monthly if f['plan'] > 0]), 1),
        'b2b_accuracy': round(np.mean([f['accuracy'] for f in b2b_fcst_monthly if f['plan'] > 0]), 1),
        # SKU 평균 정확도 (개별 동일 가중치)
        'sku_avg_accuracy_all': sku_avg_accuracy_all,
        'sku_avg_accuracy_b2b': sku_avg_accuracy_b2b,
    },
    'categories': [{'name': k, **v} for k, v in cat_sorted],
    'channels': channel_summary,
    'schedule': {
        'items': schedule_items,
        'pending_orders': pending_orders,
        'summary': {
            's95_count': len(s95_items),
            's90_count': len(s90_items),
            'order_needed': order_needed_count,
            'pending_orders_count': len(pending_orders),
            'pending_orders_qty': sum(x['qty'] for x in pending_orders),
        }
    },
}

# ── supply_gap 추가 ──
output['supply_gap'] = {
    'items': gap_items,
    'summary': {
        'shortage_count': shortage_count,
        'tight_count': tight_count,
        'ok_count': ok_count,
        'total_gap': total_gap,
    }
}

# ── 주간 변화 추적 추가 ──
output['weekly_change'] = {
    'current': alert_counts,
    'previous': prev_alerts if prev_alerts else None,
    'delta': {k: alert_counts.get(k, 0) - prev_alerts.get(k, 0) for k in alert_counts} if prev_alerts else None,
}

# ── 매출 데이터 연결 (Sales Dashboard data.json) ──
try:
    with open(SALES_JSON, encoding='utf-8') as sf:
        sales = json.load(sf)

    # 채널 매핑: inventory 채널명 → sales data 경로
    ch_sales_map = {
        '다이소': ('dom_b2b', '다이소'),
        '올리브영': ('dom_b2b', '올리브영'),
        '쿠팡': ('dom_b2c', '쿠팡'),
        '이에스(CIS)': ('ovs_indirect', 'Central Asia'),
        '동아(중국/동남아)': ('ovs_indirect', 'Greater China'),
        '에이앤리(유럽)': ('ovs_indirect', 'Southeast Asia'),  # 근사 매핑
        '힐락(일본)': ('ovs_direct', 'SEKIDO'),  # 근사 매핑
    }

    for ch_name, (segment, sales_ch) in ch_sales_map.items():
        if ch_name not in output['channels']:
            continue
        # 26년 YTD
        ytd_26 = sales.get('2026', {}).get(segment, {}).get('monthly', {}).get(sales_ch, {})
        ytd_26_total = sum(ytd_26.values())
        # 25년 동기
        ytd_25 = sales.get('2025', {}).get(segment, {}).get('monthly', {}).get(sales_ch, {})
        cur_month = sales.get('cur_month', 3)
        ytd_25_total = sum(v for k, v in ytd_25.items() if int(k) <= cur_month)
        # YoY
        yoy = ((ytd_26_total - ytd_25_total) / ytd_25_total * 100) if ytd_25_total > 0 else 0

        output['channels'][ch_name]['sales_ytd_26'] = round(ytd_26_total)
        output['channels'][ch_name]['sales_ytd_25'] = round(ytd_25_total)
        output['channels'][ch_name]['sales_yoy'] = round(yoy, 1)

    print(f"  매출 데이터 연결 완료")
except Exception as ex:
    print(f"  매출 데이터 연결 실패: {ex}")

with open(OUTPUT_JSON, 'w', encoding='utf-8') as f:
    json.dump(output, f, ensure_ascii=False, indent=None)

file_size = os.path.getsize(OUTPUT_JSON) / 1024
print(f"\n✅ inventory_data.json 생성 ({file_size:.0f}KB)")
print(f"   경로: {OUTPUT_JSON}")
print(f"   활성 SKU: {total_active}개 / 긴급: {alert_counts['critical']}개 / 과잉: {alert_counts['excess']}개")
